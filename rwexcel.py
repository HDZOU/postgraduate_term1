import re
import openpyxl
def write07Excel(path):
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = 'sheet1'
	P_list = []
	with open('test.txt','r',encoding = 'utf-8')as f:
		file = f.readlines()
		for i in range(len(file)):
			P_list.append([])
			line = file[i].split()
			for word in line:
				P_list[i].append(word)
	# value = [["名称", "价格", "出版社", "语言"],
	# 		 ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
	# 		 ["暗时间", "32.4", "人民邮电出版社", "中文"],
	# 		 ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
		for i in range(0, len(P_list)):
			for j in range(0, len(P_list[i])):
				sheet.cell(row=i+1, column=j+1, value=str(P_list[i][j]))

	wb.save(path)
	print("写入数据成功！")


def read07Excel(path):
	wb = openpyxl.load_workbook(path)
	sheet = wb.get_sheet_by_name('2007测试表')

	for row in sheet.rows:
		for cell in row:
			print(type(cell.value), "\t", end="")
		print()
def pca2text(path):
	'''将Excel中的专利正文,以及专利公司与作者数据，写入到TXT文件中'''
	wb = openpyxl.load_workbook(path)
	sheet = wb.get_sheet_by_name('发明专利')
	print('loading finished!')
	with open('patent_content.txt','w',encoding = 'utf-8') as f1:
		with open('patent_c_a.txt','w',encoding = 'utf-8') as f2:
			for row in sheet.rows:
				title = re.sub("[ |　]",'',row[1].value)
				f1.write(title+' '+row[5].value.strip()+'\n')
				f2.write(row[12].value+' '+row[13].value+'\n')
			print('writting to text')

def testinve2xls(path):
	wb = openpyxl.load_workbook(path)
	sheet = wb.get_sheet_by_name('发明专利')
	for row in sheet.rows:
		print(len(row))
def inve2xls(rpath,wpath):
	'''将Excel中的发明专利写入一个新的Excel中'''
	r_wb = openpyxl.load_workbook(rpath)
	r_sheet = r_wb.get_sheet_by_name('patent34')
	w_wb = openpyxl.Workbook()
	w_sheet = w_wb.active
	w_sheet.title = '发明专利'
	print('loading finished!')
	line = 0
	for row in r_sheet.rows:
		if row[2].value == "发明专利":
			if row[5].value == None:
				continue
			else:
				for column in range(len(row)):
					w_sheet.cell(row = line + 1,column = column + 1,value = str(row[column].value))
				line += 1
		else:
			continue
	w_wb.save(wpath)
	print('inventions have been writen to excel')

def c_a2text(path):
	 '''将Excel中的专利所属公司与专利人写入TXT文件中'''
	 wb = openpyxl.load_workbook(path)
	 sheet = wb.get_sheet_by_name('发明专利')
	 company_author_map = []
	 for row in sheet.rows:
		 c_a = []
		 c_a.append(row[12].value)
		 c_a.append(row[13].value)
		 company_author_map.append(c_a)
	 print('reading finished!')
	 with open('patent_c_a.txt','w',encoding = 'utf-8') as f:
		 for line in company_author_map:
		 	line[1] = re.sub(' ','',line[1])
		 	f.write(line[0]+' '+line[1]+'\n')
	 print('writting finished!!')
# filename = 'patent34.xlsx'
# write07Excel(file_2007)
# patent2text(filename)
# read07Excel('test.xlsx')
# c_a2text('patent34.xlsx')
if __name__ == "__main__":
	pca2text('inventions.xlsx')
	# write07Excel("新申请委员名单.xlsx")
	# testinve2xls('test.xlsx')
	# inve2xls('patent34.xlsx','test.xlsx')
	# c_a2text('inventions.xlsx')